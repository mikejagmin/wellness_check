from config import Config
import pymysql
from openpyxl import load_workbook


def import_contacts():
    con = pymysql.connect(host=Config.myhost,
                                    user=Config.myuser,
                                    password=Config.mypw,
                                    db=Config.mydb,
                                    charset='utf8mb4',
                                    cursorclass=pymysql.cursors.DictCursor)
    cur = con.cursor()
    cur.execute ('TRUNCATE TABLE contact_modify')
    wb = load_workbook('files/contact_import.xlsx', data_only=True)
    s = wb['Sheet1']
    for row in range(2,s.max_row):
        if s.cell(row,1).value is None:
            break
        location = s.cell(row,1).value
        area = s.cell(row,2).value
        country_code = s.cell(row,3).value
        phone = s.cell(row,8).value
        first_name = s.cell(row,5).value
        last_name = s.cell(row,6).value
        whatsapp = s.cell(row,7).value
        # print(location,area,country_code,phone,first_name,last_name)
        cur.execute("""INSERT INTO contact_modify (location, area, country_code, phone, whatsapp,first_name,last_name) """ + \
            """ values (%s,%s,%s,%s,%s, %s,%s);""",(location,area,country_code,phone,whatsapp,first_name,last_name))

    con.commit()
    cur.execute("""insert into contact (location,area,country_code,phone,whatsapp,first_name,last_name,active) """ + \
    """ select location, area, country_code, phone, whatsapp,first_name,last_name, 1 from contact_modify b """ + \
    """ on duplicate key update active = 1, first_name =b.first_name, last_name = b.last_name;""")
    con.commit()
    con.close()
    print('contacts imported without issue')



import_contacts()